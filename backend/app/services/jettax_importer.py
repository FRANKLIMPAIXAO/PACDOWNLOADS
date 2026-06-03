"""Importador XLSX da carteira Jettax 360 Admin.

Espera o XLSX exportado pelo Jettax com 33 colunas e 102+ linhas de empresas
(documentado em `docs/CARTEIRA-PAC-TRIBUTARIA.md`).

Idempotente por CNPJ — re-rodar atualiza campos existentes sem duplicar.

Modo `dry_run=True` valida + simula sem persistir (devolve preview).
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, asdict
from datetime import date, datetime
from io import BytesIO
from typing import Any, Literal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.empresa import Empresa


log = logging.getLogger(__name__)


# --- Mapeamentos Jettax -> PAC ---

# Coluna 14 *Tributação no Jettax: SN/MEI/LP/LR
TRIBUTACAO_MAP = {
    "SN": "Simples Nacional",
    "MEI": "MEI",
    "LP": "Lucro Presumido",
    "LR": "Lucro Real",
}

# Coluna 19 Situação Cadastral no Jettax: 1/2/3/4/8
SITUACAO_MAP = {
    1: "Nula",
    2: "Ativa",
    3: "Suspensa",
    4: "Inapta",
    8: "Baixada",
}


@dataclass(slots=True)
class ImportEmpresaResult:
    cnpj: str
    razao_social: str
    status: Literal["criada", "atualizada", "ignorada", "erro"]
    empresa_id: int | None = None
    mensagem: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class ImportXlsxResult:
    linhas_lidas: int
    criadas: int
    atualizadas: int
    ignoradas: int
    erros: int
    dry_run: bool
    detalhes: list[ImportEmpresaResult]

    def to_dict(self) -> dict[str, Any]:
        return {
            "linhas_lidas": self.linhas_lidas,
            "criadas": self.criadas,
            "atualizadas": self.atualizadas,
            "ignoradas": self.ignoradas,
            "erros": self.erros,
            "dry_run": self.dry_run,
            "detalhes": [d.to_dict() for d in self.detalhes],
        }


def _parse_date_br(value: Any) -> date | None:
    """Aceita string DD/MM/AAAA ou datetime do Excel."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None


def _str_or_none(value: Any, max_len: int | None = None) -> str | None:
    """Converte pra string limpa, ou None se vazio. Trunca em max_len se dado."""
    if value is None or value == "":
        return None
    s = str(value).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def _carregar_cidades_lookup(wb) -> dict[str, tuple[str, str]]:
    """Mapeia código IBGE → (nome_cidade, uf) da sheet `cidades`.

    Sheet Jettax tem 5573 cidades. Lookup in-memory bem rápido.
    """
    if "cidades" not in wb.sheetnames:
        return {}
    cidades: dict[str, tuple[str, str]] = {}
    for row in wb["cidades"].iter_rows(min_row=2, values_only=True):
        cod = row[0]
        if cod is None:
            continue
        nome = str(row[1] or "").strip()
        uf = str(row[2] or "").strip().upper()
        cidades[str(cod)] = (nome, uf)
    return cidades


def _montar_payload(
    row: tuple,
    cidades_map: dict[str, tuple[str, str]],
) -> dict[str, Any]:
    """Mapeia uma linha do XLSX Jettax (33 colunas) pros campos do model PAC.

    Schema de colunas (0-indexed):
        0  *CNPJ
        1  Razao Social
        2  Data de Abertura da Empresa (DD/MM/AAAA)
        3  *Inicio das Atividades na Jettax (MM/AAAA — não usado)
        4  Estado (UF)
        5  Inscricao Estadual
        6  Codigo do Municipio (IBGE)
        7  *Inscricao Municipal
        8  CEP (só dígitos)
        9  Tipo de Logradouro (R, AV, ...)
        10 Endereco
        11 Numero
        12 Complemento
        13 Bairro
        14 *Tributacao (SN/MEI/LP/LR)
        15 *Regime (1=Caixa, 2=Competencia — não mapeamos, todos são 2)
        16 Natureza Juridica
        17 CNAE
        18 Data de Vencimento do Certificado Digital
        19 Situacao Cadastral (1-8)
        20 Login Prefeitura
        21 Senha Prefeitura
        22 Codigo de Acesso Simples Nacional
        23 CPF do Responsavel do Simples Nacional
        24 Credencial (Emissor nacional)
        25 Senha (Emissor nacional)
        26 Email Fiscal
        27 Email Prevencao
        28 Telefone
        29 Integracoes
        30 Status (1=ativo, 0=inativo)
        31 Empresa Importadora (Sim/Nao)
        32 Chave Integracao Dominio
    """
    razao = _str_or_none(row[1], 255)
    if not razao:
        raise ValueError("Razao social vazia")

    # IBGE → nome cidade + UF (fallback UF da coluna 4)
    cod_municipio = _str_or_none(row[6])
    municipio_nome: str | None = None
    uf: str | None = None
    if cod_municipio and cod_municipio in cidades_map:
        municipio_nome, uf = cidades_map[cod_municipio]
    if not uf:
        uf = _str_or_none(row[4])
    if uf:
        uf = uf.upper()[:2]

    # Tributacao
    trib_raw = _str_or_none(row[14])
    regime = TRIBUTACAO_MAP.get(trib_raw.upper()) if trib_raw else None

    # Situacao cadastral (int -> nome)
    sit_raw = row[19]
    situacao = SITUACAO_MAP.get(sit_raw, "Ativa") if sit_raw is not None else "Ativa"

    # Status Jettax: 1=ativo, 0=inativo
    status_raw = row[30]
    ativo = bool(status_raw == 1 or status_raw is True or status_raw == "1")

    # CEP só dígitos, sem hífen
    cep_raw = _str_or_none(row[8])
    cep = "".join(c for c in cep_raw if c.isdigit()) if cep_raw else None
    if cep and len(cep) > 8:
        cep = cep[:8]

    payload: dict[str, Any] = {
        "razao_social": razao,
        "regime_tributario": regime,
        "tributacao": trib_raw,
        "ativo": ativo,
        # Dados cadastrais
        "inscricao_estadual": _str_or_none(row[5], 20),
        "inscricao_municipal": _str_or_none(row[7], 20),
        "data_abertura": _parse_date_br(row[2]),
        "natureza_juridica_codigo": _str_or_none(row[16], 10),
        "atividade": _str_or_none(row[17], 20),  # CNAE
        "situacao_cadastral": situacao,
        "telefone": _str_or_none(row[28], 20),
        "email_contato": _str_or_none(row[26], 120),
        # Endereço
        "cep": cep,
        "logradouro_tipo": _str_or_none(row[9], 20),
        "logradouro": _str_or_none(row[10], 200),
        "numero": _str_or_none(row[11], 20),
        "complemento": _str_or_none(row[12], 100),
        "bairro": _str_or_none(row[13], 80),
        "municipio": municipio_nome,
        "uf": uf,
        # Cert A1 — só salvamos a validade (cert .pfx em si vem em outro fluxo)
        "cert_a1_validade_ate": _parse_date_br(row[18]),
        # Credenciais Prefeitura
        "prefeitura_login": _str_or_none(row[20], 80),
        # Simples Nacional
        "simples_cpf_responsavel": _str_or_none(row[23], 11),
        # Emissor Nacional
        "emissor_nacional_login": _str_or_none(row[24], 80),
    }

    return payload


def _aplicar_credenciais_secretas(
    empresa: Empresa,
    row: tuple,
) -> None:
    """Aplica senhas cifradas. Separado pq usa setters do model."""
    senha_pref = _str_or_none(row[21])
    if senha_pref:
        empresa.set_prefeitura_senha(senha_pref)

    senha_emissor = _str_or_none(row[25])
    if senha_emissor:
        empresa.set_emissor_nacional_senha(senha_emissor)

    cod_simples = _str_or_none(row[22])
    if cod_simples:
        empresa.set_simples_codigo_acesso(cod_simples)


def importar_xlsx_jettax(
    db: Session,
    xlsx_bytes: bytes,
    *,
    dry_run: bool = False,
) -> ImportXlsxResult:
    """Importa carteira do Jettax 360 (XLSX) pro PAC.

    - Lê sheet `clientes` (102+ empresas) + cruza IBGE com sheet `cidades`.
    - UPSERT por CNPJ: cria nova ou atualiza existente.
    - `dry_run=True`: simula sem commit. Devolve preview do que faria.
    - Idempotente: re-rodar 2x = atualiza 2x mesmo registro.

    NÃO importa certificado .pfx (precisa subir separadamente — só salva
    `cert_a1_validade_ate`).

    NÃO importa `data_inicio_recebimento_nfe`/`focus_token` — esses sao
    setados pelo fluxo auto-cadastrar Focus depois.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"XLSX invalido: {type(exc).__name__}: {exc}") from exc

    if "clientes" not in wb.sheetnames:
        raise ValueError(
            f"Sheet 'clientes' nao encontrada. Sheets disponiveis: {wb.sheetnames}"
        )

    cidades_map = _carregar_cidades_lookup(wb)

    ws = wb["clientes"]
    detalhes: list[ImportEmpresaResult] = []
    criadas = atualizadas = ignoradas = erros = 0
    linhas_lidas = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        cnpj_raw = row[0]
        if cnpj_raw is None or cnpj_raw == "":
            continue  # linha vazia, pula sem contar
        linhas_lidas += 1

        # CNPJ pode vir como int (sem zeros à esquerda) ou string
        cnpj_str = str(cnpj_raw).strip()
        # Remove pontuação se houver
        cnpj_digits = "".join(c for c in cnpj_str if c.isdigit())
        if len(cnpj_digits) != 14:
            erros += 1
            detalhes.append(ImportEmpresaResult(
                cnpj=cnpj_str, razao_social=str(row[1] or ""),
                status="erro",
                mensagem=f"linha {idx}: CNPJ invalido ({len(cnpj_digits)} digitos)",
            ))
            continue
        cnpj_str = cnpj_digits

        razao_preview = str(row[1] or "").strip()[:60]

        try:
            payload = _montar_payload(row, cidades_map)
            empresa_existente = db.scalar(
                select(Empresa).where(Empresa.cnpj == cnpj_str)
            )

            if empresa_existente:
                if dry_run:
                    detalhes.append(ImportEmpresaResult(
                        cnpj=cnpj_str, razao_social=razao_preview,
                        status="atualizada", empresa_id=empresa_existente.id,
                        mensagem="(dry-run) atualizaria",
                    ))
                else:
                    for k, v in payload.items():
                        # Não sobrescreve com None — preserva dados já cadastrados
                        if v is not None:
                            setattr(empresa_existente, k, v)
                    _aplicar_credenciais_secretas(empresa_existente, row)
                    db.commit()
                    detalhes.append(ImportEmpresaResult(
                        cnpj=cnpj_str, razao_social=razao_preview,
                        status="atualizada", empresa_id=empresa_existente.id,
                    ))
                atualizadas += 1
            else:
                if dry_run:
                    detalhes.append(ImportEmpresaResult(
                        cnpj=cnpj_str, razao_social=razao_preview,
                        status="criada", mensagem="(dry-run) criaria",
                    ))
                else:
                    nova = Empresa(cnpj=cnpj_str, **payload)
                    db.add(nova)
                    db.flush()  # pra pegar o ID antes do commit
                    _aplicar_credenciais_secretas(nova, row)
                    db.commit()
                    db.refresh(nova)
                    detalhes.append(ImportEmpresaResult(
                        cnpj=cnpj_str, razao_social=razao_preview,
                        status="criada", empresa_id=nova.id,
                    ))
                criadas += 1
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            erros += 1
            detalhes.append(ImportEmpresaResult(
                cnpj=cnpj_str, razao_social=razao_preview,
                status="erro",
                mensagem=f"linha {idx}: {type(exc).__name__}: {exc}",
            ))
            log.exception("Falha ao importar linha %d (CNPJ=%s)", idx, cnpj_str)

    return ImportXlsxResult(
        linhas_lidas=linhas_lidas,
        criadas=criadas,
        atualizadas=atualizadas,
        ignoradas=ignoradas,
        erros=erros,
        dry_run=dry_run,
        detalhes=detalhes,
    )
