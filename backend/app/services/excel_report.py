from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.consulta_log import ConsultaLog
from app.models.documento_fiscal import DocumentoFiscal, TipoDocumento
from app.models.empresa import Empresa


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Brasil não tem horário de verão desde 2019 → offset fixo -03:00. Usado tanto
# pra normalizar datas do Excel quanto pros limites de competência.
BR_TZ = timezone(timedelta(hours=-3))


def _sanitize_cell(value):
    """openpyxl NÃO aceita datetime timezone-aware — o save() estoura
    `TypeError: Excel does not support timezones in datetimes`. Em produção
    (Postgres/Supabase) as colunas DateTime(timezone=True) voltam tz-aware, então
    QUALQUER célula com data derruba o relatório (500 → 'não baixou o Excel').
    Aqui convertemos pra naive na hora horária de Brasília (o horário exibido
    fica igual ao local). Em SQLite (dev) já vem naive e passa direto."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(BR_TZ).replace(tzinfo=None)
    return value


class ExcelReportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def gerar_excel_empresa(
        self,
        empresa_id: int,
        destination: Path,
        data_min: datetime | None = None,
        data_max: datetime | None = None,
    ) -> str:
        empresa = self.db.get(Empresa, empresa_id)
        if not empresa:
            raise ValueError("Empresa nao encontrada")
        stmt = select(DocumentoFiscal).where(DocumentoFiscal.empresa_id == empresa_id)
        stmt = self._aplicar_periodo(stmt, DocumentoFiscal.data_emissao, data_min, data_max)
        docs = self.db.scalars(stmt.order_by(DocumentoFiscal.data_emissao.desc())).all()
        estmt = select(ConsultaLog).where(ConsultaLog.empresa_id == empresa_id, ConsultaLog.status == "erro")
        estmt = self._aplicar_periodo(estmt, ConsultaLog.created_at, data_min, data_max)
        errors = self.db.scalars(estmt).all()
        return self._build_workbook(docs, errors, destination, empresa.razao_social)

    def gerar_excel_geral(
        self,
        destination: Path,
        data_min: datetime | None = None,
        data_max: datetime | None = None,
    ) -> str:
        stmt = self._aplicar_periodo(select(DocumentoFiscal), DocumentoFiscal.data_emissao, data_min, data_max)
        docs = self.db.scalars(stmt.order_by(DocumentoFiscal.data_emissao.desc())).all()
        estmt = self._aplicar_periodo(
            select(ConsultaLog).where(ConsultaLog.status == "erro"), ConsultaLog.created_at, data_min, data_max
        )
        errors = self.db.scalars(estmt).all()
        return self._build_workbook(docs, errors, destination, "Geral")

    @staticmethod
    def _aplicar_periodo(stmt, coluna, data_min: datetime | None, data_max: datetime | None):
        """Filtra por competência [data_min, data_max) quando informado. Sem
        limites, retorna tudo (comportamento antigo)."""
        if data_min is not None:
            stmt = stmt.where(coluna >= data_min)
        if data_max is not None:
            stmt = stmt.where(coluna < data_max)
        return stmt

    def _build_workbook(
        self,
        docs: list[DocumentoFiscal],
        errors: list[ConsultaLog],
        destination: Path,
        label: str,
    ) -> str:
        workbook = Workbook()
        resumo = workbook.active
        resumo.title = "Resumo"

        self._write_sheet(
            resumo,
            ["Cliente", "CNPJ", "Total documentos", "Total valor"],
            [[label, "", len(docs), float(sum((doc.valor_total or 0) for doc in docs))]],
        )

        self._write_document_sheet(workbook.create_sheet("NF-e"), docs, TipoDocumento.NFE)
        self._write_document_sheet(workbook.create_sheet("CT-e"), docs, TipoDocumento.CTE)
        self._write_document_sheet(workbook.create_sheet("NFS-e"), docs, TipoDocumento.NFSE)
        self._write_sheet(
            workbook.create_sheet("Erros de Consulta"),
            ["Empresa ID", "Tipo", "Status", "Mensagem", "Data"],
            [[error.empresa_id, error.tipo_documento, error.status, error.mensagem, error.created_at] for error in errors],
        )

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return str(destination)

    def _write_document_sheet(self, sheet, docs: list[DocumentoFiscal], tipo: TipoDocumento) -> None:
        filtered = [doc for doc in docs if doc.tipo_documento == tipo]
        rows = [
            [
                doc.empresa.razao_social if doc.empresa else "",
                doc.empresa.cnpj if doc.empresa else "",
                doc.tipo_documento.value,
                doc.numero,
                doc.chave_acesso,
                doc.data_emissao,
                doc.nome_emitente,
                doc.nome_destinatario,
                float(doc.valor_total or 0),
                doc.status,
                doc.xml_path,
            ]
            for doc in filtered
        ]
        self._write_sheet(
            sheet,
            [
                "Cliente",
                "CNPJ",
                "Tipo documento",
                "Numero",
                "Chave",
                "Data emissao",
                "Emitente",
                "Destinatario/Tomador",
                "Valor",
                "Status",
                "Caminho XML",
            ],
            rows,
        )

    def _write_sheet(self, sheet, headers: list[str], rows: list[list]) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for row in rows:
            # _sanitize_cell tira o tzinfo — senão o save() estoura em produção.
            sheet.append([_sanitize_cell(value) for value in row])
        sheet.auto_filter.ref = sheet.dimensions
        self._autosize(sheet)

    def _autosize(self, sheet) -> None:
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in column]
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
